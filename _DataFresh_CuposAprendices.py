import os
import win32com.client
import threading
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime
import json
import subprocess
import psutil
import shutil
import time
from PIL import Image, ImageDraw, ImageTk
import io
import openpyxl
from openpyxl.utils import get_column_letter
import requests
from io import BytesIO
import urllib.request


class ModernBancolombiaColors:
    """Paleta de colores Bancolombia profesional"""
    # Colores Bancolombia
    BANCOLOMBIA_AZUL = "#003D7A"
    BANCOLOMBIA_AMARILLO = "#FFC300"
    BANCOLOMBIA_GRIS = "#F5F5F5"
    
    # Fondos
    FONDO_PRINCIPAL = "#FFFFFF"
    FONDO_SECUNDARIO = "#F8F9FA"
    FONDO_TERCIARIO = "#E8EAED"
    
    # Primarios
    PRIMARIO = "#003D7A"
    PRIMARIO_HOVER = "#002D5A"
    PRIMARIO_CLARO = "#E3F2FD"
    
    # Secundarios
    SECUNDARIO = "#FFC300"
    SECUNDARIO_OSCURO = "#D4A000"
    SECUNDARIO_CLARO = "#FFF8E1"
    
    # Estados
    EXITO = "#28A745"
    EXITO_CLARO = "#D4EDDA"
    ALERTA = "#FFC107"
    ALERTA_CLARO = "#FFF3CD"
    ERROR = "#DC3545"
    ERROR_CLARO = "#F8D7DA"
    
    # Textos
    TEXTO_PRINCIPAL = "#1A1A1A"
    TEXTO_SECUNDARIO = "#4A4A4A"
    TEXTO_TERCIARIO = "#7A7A7A"
    
    # Acentos
    ACENTO_DORADO = "#D4A000"
    ACENTO_AZUL = "#003D7A"
    
    # Bordes
    BORDE_SUAVE = "#F5F5F5"
    BORDE_NORMAL = "#D0D0D0"


class BarraProgresoAnimada:
    """Barra de progreso animada mejorada"""
    
    def __init__(self, parent, width=400, height=12, color=ModernBancolombiaColors.PRIMARIO):
        self.parent = parent
        self.width = width
        self.height = height
        self.color = color
        self.valor = 0
        self.animando = False
        
        self.canvas = tk.Canvas(parent, width=width, height=height,
                               bg=ModernBancolombiaColors.FONDO_TERCIARIO,
                               highlightthickness=0, bd=0)
        self.rect = None
    
    def pack(self, **kwargs):
        self.canvas.pack(**kwargs)
    
    def actualizar(self, valor, animar=True):
        """Actualiza la barra con animación"""
        self.valor = min(100, max(0, valor))
        if animar:
            self._animar()
        else:
            self._dibujar()
    
    def _animar(self):
        """Animación suave de la barra"""
        if self.animando:
            return
        self.animando = True
        
        def animar_paso():
            self._dibujar()
            self.animando = False
        
        self.parent.after(50, animar_paso)
    
    def _dibujar(self):
        """Dibuja la barra con gradiente"""
        self.canvas.delete("all")
        ancho_barra = (self.valor / 100) * self.width
        
        # Fondo
        self.canvas.create_rectangle(0, 0, self.width, self.height,
                                    fill=ModernBancolombiaColors.FONDO_TERCIARIO,
                                    outline=ModernBancolombiaColors.BORDE_NORMAL)
        
        # Barra de progreso
        if ancho_barra > 0:
            self.canvas.create_rectangle(0, 0, ancho_barra, self.height,
                                        fill=self.color, outline=self.color)
            
            # Brillo superior
            self.canvas.create_rectangle(0, 0, ancho_barra * 0.3, self.height * 0.5,
                                        fill=self.color, outline="", stipple="gray50")


class NotificacionFlotante:
    """Notificación flotante emergente mejorada"""
    
    def __init__(self, padre, titulo, mensaje, tipo="info", duracion=3000):
        self.ventana = tk.Toplevel(padre)
        self.ventana.overrideredirect(True)
        self.ventana.attributes('-alpha', 0.96)
        self.ventana.attributes('-topmost', True)
        
        # Colores por tipo
        colores = {
            "info": (ModernBancolombiaColors.PRIMARIO_CLARO, ModernBancolombiaColors.PRIMARIO, "ℹ"),
            "exito": (ModernBancolombiaColors.EXITO_CLARO, ModernBancolombiaColors.EXITO, "✓"),
            "error": (ModernBancolombiaColors.ERROR_CLARO, ModernBancolombiaColors.ERROR, "✕"),
            "alerta": (ModernBancolombiaColors.ALERTA_CLARO, ModernBancolombiaColors.ALERTA, "⚠"),
        }
        
        bg_color, border_color, icono = colores.get(tipo, colores["info"])
        
        # Frame principal con sombra
        frame = tk.Frame(self.ventana, bg=bg_color, highlightthickness=2,
                        highlightbackground=border_color)
        frame.pack(fill="both", expand=True, padx=12, pady=12)
        
        # Contenido
        content = tk.Frame(frame, bg=bg_color)
        content.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Icono y título
        header = tk.Frame(content, bg=bg_color)
        header.pack(fill="x", pady=(0, 8))
        
        titulo_label = tk.Label(header, text=f"{icono} {titulo}",
                               font=("Segoe UI", 12, "bold"),
                               fg=border_color, bg=bg_color)
        titulo_label.pack(anchor="w")
        
        # Mensaje
        msg_label = tk.Label(content, text=mensaje,
                            font=("Segoe UI", 10),
                            fg=ModernBancolombiaColors.TEXTO_SECUNDARIO,
                            bg=bg_color, wraplength=350, justify="left")
        msg_label.pack(anchor="w")
        
        # Posicionar en esquina inferior derecha
        self.ventana.update_idletasks()
        x = padre.winfo_screenwidth() - 420
        y = padre.winfo_screenheight() - 200
        self.ventana.geometry(f"+{x}+{y}")
        
        # Auto-cerrar
        self.ventana.after(duracion, self.ventana.destroy)


def descargar_imagen(url, ancho=None, alto=None):
    """Descarga imagen de URL y la convierte a PhotoImage"""
    try:
        response = urllib.request.urlopen(url, timeout=5)
        img_data = response.read()
        img = Image.open(BytesIO(img_data))
        
        if ancho and alto:
            img = img.resize((ancho, alto), Image.Resampling.LANCZOS)
        
        return ImageTk.PhotoImage(img)
    except Exception as e:
        print(f"Error descargando imagen: {e}")
        return None


class CruceArchivoExcelTab:
    """Tab integrado de Cruce de Archivos Excel - VERSIÓN COMPLETA RESPONSIVE"""
    
    def __init__(self, parent, app_padre):
        self.parent = parent
        self.app_padre = app_padre
        
        # Ruta tabla de control
        self.tabla_control_path = r"C:\Users\sdcorral\OneDrive - Grupo Bancolombia\Gcia Operaciones TyC_M365 - Envio masivo de Cupos Vp\Tabla de control.xlsx"
        self.carpeta_archivos = ""
        self.archivos_seleccionados = []
        self.proceso_activo = False
        self.labels_stats_cruce = {}
        
        self._crear_interfaz()
    
    def _crear_interfaz(self):
        """Crea interfaz del tab completa responsive"""
        # Frame principal responsive
        main_frame = tk.Frame(self.parent, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Usar PanedWindow para permitir redimensionar
        paned = ttk.PanedWindow(main_frame, orient="horizontal")
        paned.pack(fill="both", expand=True)
        
        # Frame izquierdo (controles)
        left_frame = tk.Frame(paned, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        paned.add(left_frame, weight=1)
        
        panel_control = tk.Frame(left_frame, bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                                highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_NORMAL)
        panel_control.pack(fill="both", expand=True)
        
        # Canvas con scroll para el panel de controles
        canvas = tk.Canvas(panel_control, bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                          highlightthickness=0, bd=0)
        scrollbar_control = ttk.Scrollbar(panel_control, command=canvas.yview)
        scrollable_control = tk.Frame(canvas, bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        
        scrollable_control.bind("<Configure>",
                               lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        
        canvas.create_window((0, 0), window=scrollable_control, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_control.set)
        
        # Secciones de controles
        self._crear_seccion_tabla_control(scrollable_control)
        self._crear_seccion_seleccion(scrollable_control)
        self._crear_seccion_opciones_cruce(scrollable_control)
        self._crear_boton_procesar(scrollable_control)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_control.pack(side="right", fill="y")
        
        # Frame derecho (info y detalles)
        right_paned = ttk.PanedWindow(paned, orient="vertical")
        paned.add(right_paned, weight=10)
        
        # Estadísticas
        stats_frame = tk.Frame(right_paned, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        right_paned.add(stats_frame, weight=0)
        self._crear_stats_cruce(stats_frame)
        
        # Logs y detalles con tabs
        tabs_frame = tk.Frame(right_paned, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        right_paned.add(tabs_frame, weight=1)
        
        notebook = ttk.Notebook(tabs_frame)
        notebook.pack(fill="both", expand=True)
        
        tab_logs = tk.Frame(notebook, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        notebook.add(tab_logs, text="📋 Logs")
        self._crear_logs_cruce(tab_logs)
        
        tab_detalles = tk.Frame(notebook, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        notebook.add(tab_detalles, text="📊 Detalles")
        self._crear_detalles_cruce(tab_detalles)
    
    def _crear_seccion_tabla_control(self, parent):
        """Sección de tabla de control"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_TERCIARIO,
                        highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_NORMAL)
        frame.pack(fill="x", padx=12, pady=(12, 8))
        
        titulo = tk.Label(frame, text="📊 TABLA DE CONTROL",
                         font=("Segoe UI", 11, "bold"),
                         fg=ModernBancolombiaColors.PRIMARIO,
                         bg=ModernBancolombiaColors.FONDO_TERCIARIO)
        titulo.pack(anchor="w", padx=12, pady=(10, 8))
        
        self.label_tabla_ruta = tk.Label(frame, text="Ruta no configurada",
                                        font=("Segoe UI", 8),
                                        fg=ModernBancolombiaColors.ERROR,
                                        bg=ModernBancolombiaColors.FONDO_TERCIARIO,
                                        wraplength=340, justify="left")
        self.label_tabla_ruta.pack(anchor="w", padx=12, pady=(0, 8))
        
        if os.path.exists(self.tabla_control_path):
            self.label_tabla_ruta.config(text=f"✓ {self.tabla_control_path}", fg=ModernBancolombiaColors.EXITO)
        
        btn_cambiar = tk.Button(frame, text="🔧 Cambiar Ruta",
                               command=self._cambiar_ruta_tabla,
                               bg=ModernBancolombiaColors.SECUNDARIO,
                               fg=ModernBancolombiaColors.PRIMARIO,
                               font=("Segoe UI", 9, "bold"),
                               relief="flat", bd=0, padx=12, pady=8,
                               cursor="hand2", highlightthickness=0,
                               activebackground=ModernBancolombiaColors.SECUNDARIO_OSCURO)
        btn_cambiar.pack(fill="x", padx=12, pady=(0, 12))
    
    def _cambiar_ruta_tabla(self):
        """Cambia ruta de tabla de control"""
        archivo = filedialog.askopenfilename(
            title="Seleccionar tabla de control",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        
        if archivo:
            self.tabla_control_path = archivo
            self.label_tabla_ruta.config(text=f"✓ {archivo}", fg=ModernBancolombiaColors.EXITO)
            self.app_padre.add_log(f"Tabla de control: {os.path.basename(archivo)}", "importante")
    
    def _crear_seccion_seleccion(self, parent):
        """Sección de selección de carpeta"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_TERCIARIO,
                        highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_NORMAL)
        frame.pack(fill="x", padx=12, pady=(0, 8))
        
        titulo = tk.Label(frame, text="📂 CARPETA DE ARCHIVOS",
                         font=("Segoe UI", 11, "bold"),
                         fg=ModernBancolombiaColors.PRIMARIO,
                         bg=ModernBancolombiaColors.FONDO_TERCIARIO)
        titulo.pack(anchor="w", padx=12, pady=(10, 8))
        
        self.label_carpeta = tk.Label(frame, text="Sin carpeta seleccionada",
                                     font=("Segoe UI", 9),
                                     fg=ModernBancolombiaColors.ERROR,
                                     bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                                     relief="flat", bd=1, padx=10, pady=8,
                                     wraplength=340, justify="left")
        self.label_carpeta.pack(fill="x", padx=12, pady=(0, 10))
        
        btn_select = tk.Button(frame, text="📁 SELECCIONAR CARPETA",
                              command=self._seleccionar_carpeta_cruce,
                              bg=ModernBancolombiaColors.PRIMARIO,
                              fg="white",
                              font=("Segoe UI", 10, "bold"),
                              relief="flat", bd=0, padx=15, pady=10,
                              cursor="hand2", highlightthickness=0,
                              activebackground=ModernBancolombiaColors.PRIMARIO_HOVER)
        btn_select.pack(fill="x", padx=12, pady=(0, 12))
        
        # Contador
        self.label_contador_cruce = tk.Label(frame, text="0 archivo(s)",
                                            font=("Segoe UI", 8),
                                            fg=ModernBancolombiaColors.TEXTO_TERCIARIO,
                                            bg=ModernBancolombiaColors.FONDO_TERCIARIO)
        self.label_contador_cruce.pack(anchor="e", padx=12, pady=(0, 12))
    
    def _crear_seccion_opciones_cruce(self, parent):
        """Sección de opciones"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_TERCIARIO,
                        highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_NORMAL)
        frame.pack(fill="x", padx=12, pady=(0, 8))
        
        titulo = tk.Label(frame, text="⚙️ OPCIONES",
                         font=("Segoe UI", 11, "bold"),
                         fg=ModernBancolombiaColors.PRIMARIO,
                         bg=ModernBancolombiaColors.FONDO_TERCIARIO)
        titulo.pack(anchor="w", padx=12, pady=(10, 8))
        
        self.var_crear_filas = tk.BooleanVar(value=True)
        cb_crear = tk.Checkbutton(frame, text="➕ Crear nuevas filas si no existen",
                                 variable=self.var_crear_filas,
                                 bg=ModernBancolombiaColors.FONDO_TERCIARIO,
                                 fg=ModernBancolombiaColors.TEXTO_PRINCIPAL,
                                 selectcolor=ModernBancolombiaColors.FONDO_TERCIARIO,
                                 activebackground=ModernBancolombiaColors.FONDO_TERCIARIO,
                                 activeforeground=ModernBancolombiaColors.PRIMARIO,
                                 font=("Segoe UI", 8),
                                 highlightthickness=0)
        cb_crear.pack(anchor="w", padx=12, pady=3)
        
        self.var_limpiar_previo = tk.BooleanVar(value=False)
        cb_limpiar = tk.Checkbutton(frame, text="🗑️ Limpiar datos previos",
                                   variable=self.var_limpiar_previo,
                                   bg=ModernBancolombiaColors.FONDO_TERCIARIO,
                                   fg=ModernBancolombiaColors.TEXTO_PRINCIPAL,
                                   selectcolor=ModernBancolombiaColors.FONDO_TERCIARIO,
                                   activebackground=ModernBancolombiaColors.FONDO_TERCIARIO,
                                   activeforeground=ModernBancolombiaColors.PRIMARIO,
                                   font=("Segoe UI", 8),
                                   highlightthickness=0)
        cb_limpiar.pack(anchor="w", padx=12, pady=3)
        
        tk.Label(frame, text="", bg=ModernBancolombiaColors.FONDO_TERCIARIO).pack(pady=3)
    
    def _crear_boton_procesar(self, parent):
        """Botón procesar cruce"""
        btn = tk.Button(parent, text="▶ PROCESAR CRUCE",
                       command=self._procesar_cruce,
                       bg=ModernBancolombiaColors.EXITO,
                       fg="white",
                       font=("Segoe UI", 11, "bold"),
                       relief="flat", bd=0, padx=20, pady=14,
                       cursor="hand2", highlightthickness=0)
        btn.pack(fill="x", padx=12, pady=12)
    
    def _crear_stats_cruce(self, parent):
        """Estadísticas del cruce"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                        highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_NORMAL)
        frame.pack(fill="both", expand=True, padx=0, pady=0)
        
        titulo = tk.Label(frame, text="📈 ESTADÍSTICAS",
                         font=("Segoe UI", 11, "bold"),
                         fg=ModernBancolombiaColors.PRIMARIO,
                         bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        titulo.pack(anchor="w", padx=18, pady=(12, 8))
        
        # Row 1
        row1 = tk.Frame(frame, bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        row1.pack(fill="x", padx=18, pady=(0, 8))
        
        self._crear_stat_label(row1, "📄 Archivos", "0", ModernBancolombiaColors.PRIMARIO, "archivos_cruce")
        self._crear_stat_label(row1, "✓ Procesados", "0", ModernBancolombiaColors.EXITO, "procesados_cruce")
        self._crear_stat_label(row1, "⏱️ Tiempo", "00:00:00", ModernBancolombiaColors.ACENTO_DORADO, "tiempo_cruce")
        
        self.barra_progreso_cruce = BarraProgresoAnimada(frame, width=500, height=12,
                                                         color=ModernBancolombiaColors.EXITO)
        self.barra_progreso_cruce.pack(fill="x", padx=18, pady=(0, 8))
        
        self.label_porcentaje_cruce = tk.Label(frame, text="0%",
                                              font=("Segoe UI", 9, "bold"),
                                              fg=ModernBancolombiaColors.EXITO,
                                              bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        self.label_porcentaje_cruce.pack(anchor="e", padx=18, pady=(0, 8))
        
        self.label_estado_cruce = tk.Label(frame, text="Listo",
                                          font=("Segoe UI", 9),
                                          fg=ModernBancolombiaColors.EXITO,
                                          bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        self.label_estado_cruce.pack(anchor="w", padx=18, pady=(0, 12))
    
    def _crear_stat_label(self, parent, titulo, valor, color, key):
        """Crea etiqueta de estadística"""
        card = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_PRINCIPAL,
                       highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_SUAVE)
        card.pack(side="left", fill="both", expand=True, padx=(0, 6))
        
        label_titulo = tk.Label(card, text=titulo,
                               font=("Segoe UI", 8),
                               fg=ModernBancolombiaColors.TEXTO_TERCIARIO,
                               bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        label_titulo.pack(padx=10, pady=(8, 3))
        
        label_valor = tk.Label(card, text=valor,
                              font=("Segoe UI", 12, "bold"),
                              fg=color,
                              bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        label_valor.pack(padx=10, pady=(3, 8))
        
        self.labels_stats_cruce[key] = label_valor
    
    def _crear_logs_cruce(self, parent):
        """Sección de logs"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        frame.pack(fill="both", expand=True, padx=8, pady=8)
        
        self.txt_log_cruce = tk.Text(frame,
                                    bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                                    fg=ModernBancolombiaColors.TEXTO_PRINCIPAL,
                                    font=("Consolas", 9),
                                    state='disabled',
                                    borderwidth=1,
                                    relief="flat",
                                    insertbackground=ModernBancolombiaColors.PRIMARIO,
                                    selectbackground=ModernBancolombiaColors.PRIMARIO_CLARO)
        self.txt_log_cruce.pack(side="left", fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(frame, command=self.txt_log_cruce.yview)
        scrollbar.pack(side="right", fill="y")
        self.txt_log_cruce.config(yscrollcommand=scrollbar.set)
        
        # Tags de colores
        self.txt_log_cruce.tag_config("info", foreground=ModernBancolombiaColors.PRIMARIO)
        self.txt_log_cruce.tag_config("exito", foreground=ModernBancolombiaColors.EXITO,
                                     font=("Consolas", 9, "bold"))
        self.txt_log_cruce.tag_config("error", foreground=ModernBancolombiaColors.ERROR,
                                     font=("Consolas", 9, "bold"))
        self.txt_log_cruce.tag_config("advertencia", foreground=ModernBancolombiaColors.ALERTA)
    
    def _crear_detalles_cruce(self, parent):
        """Tabla de detalles del cruce"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        frame.pack(fill="both", expand=True, padx=8, pady=8)
        
        columns = ("# ", "Archivo", "C2", "C4", "Hora")
        self.tree_cruce = ttk.Treeview(frame, columns=columns, height=15)
        self.tree_cruce.column("#0", width=0, stretch=tk.NO)
        self.tree_cruce.column("# ", anchor=tk.CENTER, width=35)
        self.tree_cruce.column("Archivo", anchor=tk.W, width=250)
        self.tree_cruce.column("C2", anchor=tk.CENTER, width=80)
        self.tree_cruce.column("C4", anchor=tk.CENTER, width=80)
        self.tree_cruce.column("Hora", anchor=tk.CENTER, width=80)
        
        self.tree_cruce.heading("#0", text="")
        self.tree_cruce.heading("# ", text="#")
        self.tree_cruce.heading("Archivo", text="Archivo")
        self.tree_cruce.heading("C2", text="Valor C2")
        self.tree_cruce.heading("C4", text="Valor C4")
        self.tree_cruce.heading("Hora", text="Hora")
        
        self.tree_cruce.pack(fill="both", expand=True)
    
    def add_log_cruce(self, mensaje, tipo="info"):
        """Agrega log al tab"""
        ahora = datetime.now().strftime("%H:%M:%S")
        iconos = {
            "info": "ℹ",
            "exito": "✓",
            "error": "✕",
            "advertencia": "⚠",
        }
        
        icono = iconos.get(tipo, "•")
        texto = f"[{ahora}] {icono} {mensaje}\n"
        
        self.txt_log_cruce.config(state='normal')
        self.txt_log_cruce.insert(tk.END, texto, tipo)
        self.txt_log_cruce.see(tk.END)
        self.txt_log_cruce.config(state='disabled')
    
    def _seleccionar_carpeta_cruce(self):
        """Selecciona carpeta"""
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta con archivos Excel")
        
        if not carpeta:
            return
        
        self.carpeta_archivos = carpeta
        self.archivos_seleccionados = []
        
        # Buscar archivos Excel
        for archivo in os.listdir(carpeta):
            if archivo.endswith(('.xlsx', '.xls', '.xlsm')):
                self.archivos_seleccionados.append(os.path.join(carpeta, archivo))
        
        cantidad = len(self.archivos_seleccionados)
        self.label_carpeta.config(text=f"✓ {carpeta}", fg=ModernBancolombiaColors.EXITO)
        self.label_contador_cruce.config(text=f"✓ {cantidad} archivo(s)")
        self.add_log_cruce(f"Carpeta seleccionada: {os.path.basename(carpeta)}", "exito")
    
    def _procesar_cruce(self):
        """Inicia procesamiento"""
        if not self.archivos_seleccionados:
            messagebox.showwarning("Advertencia", "Debes seleccionar una carpeta primero")
            return
        
        if not os.path.exists(self.tabla_control_path):
            messagebox.showerror("Error", f"No se encontró tabla de control:\n{self.tabla_control_path}")
            return
        
        self.proceso_activo = True
        self.barra_progreso_cruce.actualizar(0)
        self.label_estado_cruce.config(text="Procesando...", fg=ModernBancolombiaColors.ALERTA)
        
        hilo = threading.Thread(target=self._ejecutar_cruce, daemon=True)
        hilo.start()
    
    def _ejecutar_cruce(self):
        """Ejecuta el cruce"""
        try:
            self.add_log_cruce("=" * 80, "info")
            self.add_log_cruce(f"INICIANDO CRUCE - {len(self.archivos_seleccionados)} archivo(s)", "exito")
            self.add_log_cruce("=" * 80, "info")
            
            datos_cruce = {}
            total_archivos = len(self.archivos_seleccionados)
            
            # Leer datos de cada archivo
            for idx, archivo in enumerate(self.archivos_seleccionados):
                try:
                    nombre_archivo = os.path.basename(archivo)
                    nombre_sin_ext = os.path.splitext(nombre_archivo)[0]
                    
                    self.add_log_cruce(f"[{idx + 1}/{total_archivos}] Leyendo: {nombre_sin_ext}...", "info")
                    
                    # Abrir archivo
                    wb = openpyxl.load_workbook(archivo, data_only=True)
                    ws = wb.active
                    
                    # Obtener valores de C2 y C4
                    valor_c2 = ws['C2'].value
                    valor_c4 = ws['C4'].value
                    
                    valor_c2_texto = str(valor_c2) if valor_c2 is not None else ""
                    valor_c4_texto = str(valor_c4) if valor_c4 is not None else ""
                    
                    datos_cruce[nombre_sin_ext] = {
                        'c2': valor_c2_texto,
                        'c4': valor_c4_texto,
                        'archivo': nombre_archivo
                    }
                    
                    wb.close()
                    
                    # Actualizar progreso
                    progreso = int((idx + 1) / total_archivos * 50)
                    self.barra_progreso_cruce.actualizar(progreso)
                    self.label_porcentaje_cruce.config(text=f"{progreso}%")
                    
                except Exception as e:
                    self.add_log_cruce(f"Error leyendo {nombre_archivo}: {str(e)[:50]}", "error")
            
            # Actualizar tabla de control
            self.add_log_cruce("Actualizando tabla de control...", "info")
            self._actualizar_tabla_control(datos_cruce)
            
            # Actualizar tabla de detalles
            self.tree_cruce.delete(*self.tree_cruce.get_children())
            for idx, (nombre, datos) in enumerate(datos_cruce.items(), 1):
                self.tree_cruce.insert("", "end", values=(
                    str(idx),
                    nombre[:40] + "..." if len(nombre) > 40 else nombre,
                    datos['c2'][:20] if len(datos['c2']) > 20 else datos['c2'],
                    datos['c4'][:20] if len(datos['c4']) > 20 else datos['c4'],
                    datetime.now().strftime("%H:%M:%S")
                ))
            
            # Finalizar
            self.barra_progreso_cruce.actualizar(100)
            self.label_porcentaje_cruce.config(text="100%")
            self.labels_stats_cruce['archivos_cruce'].config(text=str(total_archivos))
            self.labels_stats_cruce['procesados_cruce'].config(text=str(len(datos_cruce)))
            self.label_estado_cruce.config(text="✓ Completado", fg=ModernBancolombiaColors.EXITO)
            
            self.add_log_cruce("=" * 80, "info")
            self.add_log_cruce(f"✓ CRUCE COMPLETADO - {len(datos_cruce)} archivo(s) procesado(s)", "exito")
            self.add_log_cruce("=" * 80, "info")
            
            NotificacionFlotante(self.parent.master if hasattr(self.parent, 'master') else self.parent,
                                "✓ Cruce Completado",
                                f"Se procesaron {len(datos_cruce)} archivo(s)\n"
                                f"Tabla de control actualizada", "exito", 3000)
            
        except Exception as e:
            self.add_log_cruce(f"ERROR: {str(e)[:100]}", "error")
            self.label_estado_cruce.config(text="✕ Error", fg=ModernBancolombiaColors.ERROR)
            messagebox.showerror("Error", f"Error en el cruce:\n{str(e)}")
        
        finally:
            self.proceso_activo = False
    
    def _actualizar_tabla_control(self, datos_cruce):
        """Actualiza tabla de control"""
        try:
            wb = openpyxl.load_workbook(self.tabla_control_path)
            ws = wb.active
            
            fila_inicio = 2
            fila_actual = fila_inicio
            
            for nombre_sin_ext, datos in datos_cruce.items():
                encontrado = False
                
                # Buscar coincidencia en columna F
                for fila in range(fila_inicio, ws.max_row + 1):
                    celda_f = ws[f'F{fila}'].value
                    
                    if celda_f and str(celda_f).strip() == nombre_sin_ext.strip():
                        ws[f'G{fila}'].value = datos['c2']
                        ws[f'E{fila}'].value = datos['c4']
                        encontrado = True
                        break
                
                # Si no existe y la opción está habilitada
                if not encontrado and self.var_crear_filas.get():
                    ws[f'F{fila_actual}'].value = nombre_sin_ext
                    ws[f'G{fila_actual}'].value = datos['c2']
                    ws[f'E{fila_actual}'].value = datos['c4']
                    fila_actual += 1
            
            wb.save(self.tabla_control_path)
            wb.close()
            
            self.add_log_cruce(f"✓ Tabla de control actualizada", "exito")
            
        except Exception as e:
            self.add_log_cruce(f"Error actualizando tabla: {str(e)}", "error")
            raise


class DataRefreshProV6:
    def __init__(self, root):
        self.root = root
        self.root.title("DataRefresh Cupos Aprendices - Bancolombia")
        self.root.iconbitmap(default="")  # Sin icono por defecto
        
        # Obtener dimensiones de pantalla
        ancho_pantalla = self.root.winfo_screenwidth()
        alto_pantalla = self.root.winfo_screenheight()
        
        # Usar 98% de la pantalla
        ancho_ventana = int(ancho_pantalla * 0.98)
        alto_ventana = int(alto_pantalla * 0.98)
        
        self.root.geometry(f"{ancho_ventana}x{alto_ventana}")
        self.root.configure(bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        self.root.resizable(True, True)
        
        # Carpeta documentos
        self.DOCUMENTOS = os.path.expanduser("~\\Documents")
        self.archivo_actual = self.DOCUMENTOS
        self.archivos_seleccionados = []
        self.proceso_activo = False
        
        # Ruta de logs
        self.ruta_logs = r"C:\Users\sdcorral\OneDrive - Grupo Bancolombia\Gcia Operaciones TyC_M365 - Envio masivo de Cupos Vp\Logs"
        
        # Variables de tiempo real
        self.tiempo_inicio_general = None
        self.tiempo_inicio_archivo = None
        self.archivo_actual_procesando = ""
        
        # Estadísticas
        self.estadisticas = {
            'total': 0,
            'procesados': 0,
            'exitosos': 0,
            'fallidos': 0,
            'tiempo_total': 0,
            'archivos_detalles': []
        }
        
        # Logs para exportar
        self.logs_procesamiento = []
        
        self.labels_stats = {}
        self.cargar_configuracion()
        self._configurar_estilos()
        self._crear_interfaz()
        self._iniciar_actualizacion_tiempo()
        self.add_log("DataRefresh Cupos Aprendices iniciado ✓", "exito")

    def _configurar_estilos(self):
        """Configura estilos de ttk"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Progressbar
        style.configure("Custom.Horizontal.TProgressbar",
                       background=ModernBancolombiaColors.PRIMARIO,
                       troughcolor=ModernBancolombiaColors.FONDO_TERCIARIO,
                       bordercolor=ModernBancolombiaColors.BORDE_NORMAL,
                       lightcolor=ModernBancolombiaColors.PRIMARIO,
                       darkcolor=ModernBancolombiaColors.PRIMARIO_HOVER,
                       thickness=12)
        
        # Treeview
        style.configure("Treeview",
                       background=ModernBancolombiaColors.FONDO_SECUNDARIO,
                       foreground=ModernBancolombiaColors.TEXTO_PRINCIPAL,
                       fieldbackground=ModernBancolombiaColors.FONDO_SECUNDARIO,
                       borderwidth=1)
        style.configure("Treeview.Heading",
                       background=ModernBancolombiaColors.FONDO_TERCIARIO,
                       foreground=ModernBancolombiaColors.TEXTO_PRINCIPAL,
                       borderwidth=1)
        style.map("Treeview", background=[('selected', ModernBancolombiaColors.PRIMARIO_CLARO)])

    def _crear_interfaz(self):
        """Crea interfaz completa con diseño Bancolombia"""
        # Container principal
        main_container = tk.Frame(self.root, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        main_container.pack(fill="both", expand=True)
        
        # Header
        self._crear_header(main_container)
        
        # Contenido responsive con tabs
        self._crear_contenido_tabs(main_container)

    def _crear_header(self, parent):
        """Header con branding Bancolombia"""
        header = tk.Frame(parent, bg=ModernBancolombiaColors.PRIMARIO,
                         highlightthickness=0, relief="flat")
        header.pack(fill="x")
        
        content = tk.Frame(header, bg=ModernBancolombiaColors.PRIMARIO)
        content.pack(fill="both", expand=True, padx=25, pady=15)
        
        # Contenedor con logo y texto
        top_row = tk.Frame(content, bg=ModernBancolombiaColors.PRIMARIO)
        top_row.pack(fill="x", pady=(0, 10))
        
        # Logo (intentar descargar)
        try:
            logo_url = "https://sites.sharepoint.com/teams/co-csi/SiteAssets/__sitelogo__logo-banco-isotipo-negro.png"
            logo_img = descargar_imagen(logo_url, 40, 40)
            if logo_img:
                logo_label = tk.Label(top_row, image=logo_img, bg=ModernBancolombiaColors.PRIMARIO)
                logo_label.image = logo_img  # Mantener referencia
                logo_label.pack(side="left", padx=(0, 15))
        except:
            pass
        
        # Título
        titulo_frame = tk.Frame(content, bg=ModernBancolombiaColors.PRIMARIO)
        titulo_frame.pack(anchor="w")
        
        titulo = tk.Label(titulo_frame, text="⚡ DataRefresh Cupos Aprendices",
                         font=("Segoe UI", 24, "bold"),
                         fg="white",
                         bg=ModernBancolombiaColors.PRIMARIO)
        titulo.pack(anchor="w")
        
        # Subtítulo
        subtitulo = tk.Label(content, 
                            text="Sincronización de insumos para cupos de aprendices | Actualización de Conexiones en tiempo real | Desarrollado por: Sdcorral",
                            font=("Segoe UI", 10),
                            fg=ModernBancolombiaColors.SECUNDARIO,
                            bg=ModernBancolombiaColors.PRIMARIO)
        subtitulo.pack(anchor="w", pady=(5, 0))

    def _crear_contenido_tabs(self, parent):
        """Crea tabs principales con contenido responsive"""
        notebook = ttk.Notebook(parent)
        notebook.pack(fill="both", expand=True, padx=0, pady=0)
        
        # TAB 1: PROCESAMIENTO PRINCIPAL
        tab_principal = tk.Frame(notebook, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        notebook.add(tab_principal, text="⚡ Procesamiento Principal")
        self._crear_tab_principal(tab_principal)
        
        # TAB 2: CRUCE DE ARCHIVOS
        tab_cruce = tk.Frame(notebook, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        notebook.add(tab_cruce, text="🔄 Cruce de Archivos")
        self.tab_cruce = CruceArchivoExcelTab(tab_cruce, self)
        
        # TAB 3: SISTEMA
        tab_sistema = tk.Frame(notebook, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        notebook.add(tab_sistema, text="🖥️ Sistema")
        self._crear_tab_sistema(tab_sistema)

    def _crear_tab_principal(self, parent):
        """Tab principal con layout responsive"""
        # PanedWindow horizontal para 2 columnas
        paned = ttk.PanedWindow(parent, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=0, pady=0)
        
        # Columna izquierda - Controles
        left_panel = tk.Frame(paned, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        paned.add(left_panel, weight=1)
        
        panel_control = tk.Frame(left_panel, bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                                highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_NORMAL)
        panel_control.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Canvas con scroll
        canvas = tk.Canvas(panel_control, bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                          highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(panel_control, command=canvas.yview)
        scrollable = tk.Frame(canvas, bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        
        scrollable.bind("<Configure>",
                       lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        self._crear_seccion_archivos(scrollable)
        self._crear_seccion_opciones(scrollable)
        self._crear_seccion_botones(scrollable)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Columna derecha - Información responsive
        right_paned = ttk.PanedWindow(paned, orient="vertical")
        paned.add(right_paned, weight=10)
        
        # Estadísticas
        stats_frame = tk.Frame(right_paned, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        right_paned.add(stats_frame, weight=0)
        self._crear_panel_estadisticas(stats_frame)
        
        # Tabs con logs y detalles
        tabs_frame = tk.Frame(right_paned, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        right_paned.add(tabs_frame, weight=1)
        
        notebook_interno = ttk.Notebook(tabs_frame)
        notebook_interno.pack(fill="both", expand=True, padx=15, pady=15)
        
        tab_logs = tk.Frame(notebook_interno, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        notebook_interno.add(tab_logs, text="📋 Logs")
        self._crear_seccion_logs(tab_logs)
        
        tab_detalles = tk.Frame(notebook_interno, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        notebook_interno.add(tab_detalles, text="📊 Detalles")
        self._crear_seccion_detalles(tab_detalles)

    def _crear_seccion_archivos(self, parent):
        """Sección de archivos"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_TERCIARIO,
                        highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_NORMAL)
        frame.pack(fill="x", padx=12, pady=(12, 8))
        
        titulo = tk.Label(frame, text="📂 ARCHIVOS",
                         font=("Segoe UI", 11, "bold"),
                         fg=ModernBancolombiaColors.PRIMARIO,
                         bg=ModernBancolombiaColors.FONDO_TERCIARIO)
        titulo.pack(anchor="w", padx=12, pady=(10, 8))
        
        self.label_archivos_info = tk.Label(frame, text="0 archivo(s) seleccionado(s)",
                                           font=("Segoe UI", 10, "bold"),
                                           fg=ModernBancolombiaColors.SECUNDARIO,
                                           bg=ModernBancolombiaColors.FONDO_TERCIARIO)
        self.label_archivos_info.pack(anchor="w", padx=12, pady=(0, 10))
        
        # Botones
        btn_frame = tk.Frame(frame, bg=ModernBancolombiaColors.FONDO_TERCIARIO)
        btn_frame.pack(fill="x", padx=12, pady=(0, 12))
        
        btn1 = tk.Button(btn_frame, text="📄 Seleccionar Archivos",
                        command=self._seleccionar_archivos,
                        bg=ModernBancolombiaColors.SECUNDARIO,
                        fg=ModernBancolombiaColors.PRIMARIO,
                        font=("Segoe UI", 9, "bold"),
                        relief="flat", bd=0, padx=12, pady=10,
                        cursor="hand2", highlightthickness=0,
                        activebackground=ModernBancolombiaColors.SECUNDARIO_OSCURO)
        btn1.pack(side="left", fill="x", expand=True, padx=(0, 6))
        
        btn2 = tk.Button(btn_frame, text="📁 Seleccionar Carpeta",
                        command=self._seleccionar_carpeta,
                        bg=ModernBancolombiaColors.PRIMARIO,
                        fg="white",
                        font=("Segoe UI", 9, "bold"),
                        relief="flat", bd=0, padx=12, pady=10,
                        cursor="hand2", highlightthickness=0,
                        activebackground=ModernBancolombiaColors.PRIMARIO_HOVER)
        btn2.pack(side="left", fill="x", expand=True, padx=(6, 0))

    def _crear_seccion_opciones(self, parent):
        """Sección de opciones"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_TERCIARIO,
                        highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_NORMAL)
        frame.pack(fill="x", padx=12, pady=(0, 8))
        
        titulo = tk.Label(frame, text="⚙️ OPCIONES",
                         font=("Segoe UI", 11, "bold"),
                         fg=ModernBancolombiaColors.PRIMARIO,
                         bg=ModernBancolombiaColors.FONDO_TERCIARIO)
        titulo.pack(anchor="w", padx=12, pady=(10, 8))
        
        opciones = [
            ("💾 Guardar automáticamente", True, "var_guardar"),
            ("🔄 Recalcular fórmulas", False, "var_calcular"),
            ("🔗 Actualizar conexiones", True, "var_conexiones"),
            ("⚙️ Ejecutar macros", False, "var_macros"),
            ("📢 Notificaciones", True, "var_notificaciones"),
            ("📂 Abrir archivo al finalizar", False, "var_abrir_archivo"),
        ]
        
        for texto, default, var_name in opciones:
            var = tk.BooleanVar(value=default)
            setattr(self, var_name, var)
            
            cb = tk.Checkbutton(frame, text=texto,
                              variable=var,
                              bg=ModernBancolombiaColors.FONDO_TERCIARIO,
                              fg=ModernBancolombiaColors.TEXTO_PRINCIPAL,
                              selectcolor=ModernBancolombiaColors.FONDO_TERCIARIO,
                              activebackground=ModernBancolombiaColors.FONDO_TERCIARIO,
                              activeforeground=ModernBancolombiaColors.PRIMARIO,
                              font=("Segoe UI", 8),
                              highlightthickness=0)
            cb.pack(anchor="w", padx=12, pady=3)
        
        tk.Label(frame, text="", bg=ModernBancolombiaColors.FONDO_TERCIARIO).pack(pady=3)

    def _crear_seccion_botones(self, parent):
        """Botones principales"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_TERCIARIO)
        frame.pack(fill="x", padx=12, pady=(0, 12))
        
        self.btn_iniciar = tk.Button(frame, text="▶ INICIAR PROCESAMIENTO",
                                    command=self._iniciar_proceso,
                                    bg=ModernBancolombiaColors.EXITO,
                                    fg="white",
                                    font=("Segoe UI", 10, "bold"),
                                    relief="flat", bd=0, padx=15, pady=14,
                                    cursor="hand2", highlightthickness=0)
        self.btn_iniciar.pack(fill="x", pady=(0, 8))
        
        # Botón Activar Automate (exportar logs)
        btn_automate = tk.Button(frame, text="💾 ACTIVAR AUTOMATE",
                                command=self._exportar_logs,
                                bg=ModernBancolombiaColors.SECUNDARIO,
                                fg=ModernBancolombiaColors.PRIMARIO,
                                font=("Segoe UI", 10, "bold"),
                                relief="flat", bd=0, padx=15, pady=12,
                                cursor="hand2", highlightthickness=0,
                                activebackground=ModernBancolombiaColors.SECUNDARIO_OSCURO)
        btn_automate.pack(fill="x")

    def _crear_panel_estadisticas(self, parent):
        """Panel de estadísticas completo"""
        panel = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                        highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_NORMAL)
        panel.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Título
        titulo = tk.Label(panel, text="📈 ESTADÍSTICAS EN TIEMPO REAL",
                         font=("Segoe UI", 11, "bold"),
                         fg=ModernBancolombiaColors.PRIMARIO,
                         bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        titulo.pack(anchor="w", padx=18, pady=(12, 8))
        
        # Archivo actual
        archivo_frame = tk.Frame(panel, bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        archivo_frame.pack(fill="x", padx=18, pady=(0, 12))
        
        label_archivo_titulo = tk.Label(archivo_frame, text="Archivo Actual:",
                                       font=("Segoe UI", 9, "bold"),
                                       fg=ModernBancolombiaColors.TEXTO_TERCIARIO,
                                       bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        label_archivo_titulo.pack(anchor="w")
        
        self.label_archivo_actual = tk.Label(archivo_frame, text="--",
                                            font=("Segoe UI", 10, "bold"),
                                            fg=ModernBancolombiaColors.PRIMARIO,
                                            bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                                            wraplength=500)
        self.label_archivo_actual.pack(anchor="w")
        
        # Grid de estadísticas
        stats_frame = tk.Frame(panel, bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        stats_frame.pack(fill="x", padx=18, pady=(0, 12))
        
        # Row 1
        row1 = tk.Frame(stats_frame, bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        row1.pack(fill="x", pady=(0, 8))
        
        self._crear_stat_card(row1, "📊 Total", "0", ModernBancolombiaColors.PRIMARIO, "total")
        self._crear_stat_card(row1, "⏳ Procesados", "0", ModernBancolombiaColors.SECUNDARIO, "procesados")
        self._crear_stat_card(row1, "✓ Exitosos", "0", ModernBancolombiaColors.EXITO, "exitosos")
        
        # Row 2
        row2 = tk.Frame(stats_frame, bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        row2.pack(fill="x", pady=(0, 8))
        
        self._crear_stat_card(row2, "✕ Fallidos", "0", ModernBancolombiaColors.ERROR, "fallidos")
        self._crear_stat_card(row2, "⏱️ Tiempo Total", "00:00:00", ModernBancolombiaColors.ACENTO_DORADO, "tiempo_total")
        self._crear_stat_card(row2, "⏱️ Archivo", "00:00:00", ModernBancolombiaColors.ACENTO_AZUL, "tiempo_archivo")
        
        # Barra de progreso
        prog_frame = tk.Frame(panel, bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        prog_frame.pack(fill="x", padx=18, pady=(0, 12))
        
        label_prog = tk.Label(prog_frame, text="Progreso General:",
                             font=("Segoe UI", 9, "bold"),
                             fg=ModernBancolombiaColors.TEXTO_TERCIARIO,
                             bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        label_prog.pack(anchor="w", pady=(0, 6))
        
        self.barra_progreso = BarraProgresoAnimada(prog_frame, width=540, height=14,
                                                   color=ModernBancolombiaColors.PRIMARIO)
        self.barra_progreso.pack(fill="x")
        
        # Porcentaje
        self.label_porcentaje = tk.Label(prog_frame, text="0%",
                                        font=("Segoe UI", 9, "bold"),
                                        fg=ModernBancolombiaColors.PRIMARIO,
                                        bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        self.label_porcentaje.pack(anchor="e", pady=(4, 0))

    def _crear_stat_card(self, parent, titulo, valor_inicial, color, key):
        """Crea tarjeta de estadística"""
        card = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_PRINCIPAL,
                       highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_SUAVE)
        card.pack(side="left", fill="both", expand=True, padx=(0, 6))
        
        label_titulo = tk.Label(card, text=titulo,
                               font=("Segoe UI", 8),
                               fg=ModernBancolombiaColors.TEXTO_TERCIARIO,
                               bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        label_titulo.pack(padx=10, pady=(8, 3))
        
        label_valor = tk.Label(card, text=valor_inicial,
                              font=("Segoe UI", 14, "bold"),
                              fg=color,
                              bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        label_valor.pack(padx=10, pady=(3, 8))
        
        self.labels_stats[key] = label_valor

    def _crear_seccion_logs(self, parent):
        """Sección de logs"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        frame.pack(fill="both", expand=True, padx=8, pady=8)
        
        self.txt_log = tk.Text(frame,
                              bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                              fg=ModernBancolombiaColors.TEXTO_PRINCIPAL,
                              font=("Consolas", 9),
                              state='disabled',
                              borderwidth=1,
                              relief="flat",
                              insertbackground=ModernBancolombiaColors.PRIMARIO,
                              selectbackground=ModernBancolombiaColors.PRIMARIO_CLARO)
        self.txt_log.pack(side="left", fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(frame, command=self.txt_log.yview)
        scrollbar.pack(side="right", fill="y")
        self.txt_log.config(yscrollcommand=scrollbar.set)
        
        # Tags de colores
        self.txt_log.tag_config("info", foreground=ModernBancolombiaColors.PRIMARIO)
        self.txt_log.tag_config("exito", foreground=ModernBancolombiaColors.EXITO, font=("Consolas", 9, "bold"))
        self.txt_log.tag_config("error", foreground=ModernBancolombiaColors.ERROR, font=("Consolas", 9, "bold"))
        self.txt_log.tag_config("advertencia", foreground=ModernBancolombiaColors.ALERTA)
        self.txt_log.tag_config("importante", foreground=ModernBancolombiaColors.ACENTO_AZUL,
                               font=("Consolas", 10, "bold"))
        self.txt_log.tag_config("archivo", foreground=ModernBancolombiaColors.PRIMARIO,
                               font=("Consolas", 9, "bold"))

    def _crear_seccion_detalles(self, parent):
        """Tabla de detalles"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        frame.pack(fill="both", expand=True, padx=8, pady=8)
        
        columns = ("# ", "Archivo", "Estado", "Duración", "Hora")
        tree = ttk.Treeview(frame, columns=columns, height=20)
        tree.column("#0", width=0, stretch=tk.NO)
        tree.column("# ", anchor=tk.CENTER, width=35)
        tree.column("Archivo", anchor=tk.W, width=300)
        tree.column("Estado", anchor=tk.CENTER, width=100)
        tree.column("Duración", anchor=tk.CENTER, width=90)
        tree.column("Hora", anchor=tk.CENTER, width=100)
        
        tree.heading("#0", text="")
        tree.heading("# ", text="#")
        tree.heading("Archivo", text="Archivo")
        tree.heading("Estado", text="Estado")
        tree.heading("Duración", text="Duración (s)")
        tree.heading("Hora", text="Hora")
        
        self.tree_detalles = tree
        tree.pack(fill="both", expand=True)

    def _crear_tab_sistema(self, parent):
        """Tab de información del sistema"""
        frame = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        titulo = tk.Label(frame, text="🖥️ Recursos del Sistema",
                         font=("Segoe UI", 16, "bold"),
                         fg=ModernBancolombiaColors.PRIMARIO,
                         bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        titulo.pack(anchor="w", pady=(0, 20))
        
        # Grid de información
        info_grid = tk.Frame(frame, bg=ModernBancolombiaColors.FONDO_PRINCIPAL)
        info_grid.pack(fill="x", pady=10)
        
        # CPU
        try:
            cpu_percent = psutil.cpu_percent()
            cpu_info = f"💻 CPU: {psutil.cpu_count()} núcleos | {cpu_percent}% en uso"
            cpu_color = ModernBancolombiaColors.EXITO if cpu_percent < 50 else ModernBancolombiaColors.ALERTA if cpu_percent < 80 else ModernBancolombiaColors.ERROR
            self._crear_info_card(info_grid, "CPU", cpu_info, cpu_color)
        except:
            pass
        
        # RAM
        try:
            mem = psutil.virtual_memory()
            mem_info = f"💾 RAM: {mem.percent}% en uso ({mem.used // (1024**3)}GB / {mem.total // (1024**3)}GB)"
            mem_color = ModernBancolombiaColors.EXITO if mem.percent < 50 else ModernBancolombiaColors.ALERTA if mem.percent < 80 else ModernBancolombiaColors.ERROR
            self._crear_info_card(info_grid, "Memoria", mem_info, mem_color)
        except:
            pass
        
        # Disco
        try:
            disk = shutil.disk_usage("C:")
            disk_percent = (disk.used / disk.total) * 100
            disk_info = f"💿 Disco C: {disk_percent:.1f}% en uso ({disk.free // (1024**3)}GB libres)"
            disk_color = ModernBancolombiaColors.EXITO if disk_percent < 70 else ModernBancolombiaColors.ALERTA if disk_percent < 90 else ModernBancolombiaColors.ERROR
            self._crear_info_card(info_grid, "Disco", disk_info, disk_color)
        except:
            pass
        
        # Usuario
        try:
            usuario = os.getenv('USERNAME')
            user_info = f"👤 Usuario: {usuario}"
            self._crear_info_card(info_grid, "Usuario", user_info, ModernBancolombiaColors.PRIMARIO)
        except:
            pass
        
        # Documentos
        try:
            doc_info = f"📁 Documentos: {self.DOCUMENTOS}"
            self._crear_info_card(info_grid, "Carpeta", doc_info, ModernBancolombiaColors.SECUNDARIO)
        except:
            pass

    def _crear_info_card(self, parent, titulo, info, color):
        """Crea tarjeta de información"""
        card = tk.Frame(parent, bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                       highlightthickness=1, highlightbackground=ModernBancolombiaColors.BORDE_NORMAL)
        card.pack(fill="x", padx=10, pady=8)
        
        title_label = tk.Label(card, text=titulo,
                              font=("Segoe UI", 11, "bold"),
                              fg=color,
                              bg=ModernBancolombiaColors.FONDO_SECUNDARIO)
        title_label.pack(anchor="w", padx=15, pady=(12, 6))
        
        info_label = tk.Label(card, text=info,
                             font=("Segoe UI", 10),
                             fg=ModernBancolombiaColors.TEXTO_SECUNDARIO,
                             bg=ModernBancolombiaColors.FONDO_SECUNDARIO,
                             wraplength=700, justify="left")
        info_label.pack(anchor="w", padx=15, pady=(0, 12))

    def add_log(self, mensaje, tipo="info"):
        """Agrega log"""
        ahora = datetime.now().strftime("%H:%M:%S")
        iconos = {
            "info": "ℹ",
            "exito": "✓",
            "error": "✕",
            "advertencia": "⚠",
            "importante": "●",
            "archivo": "📄"
        }
        
        icono = iconos.get(tipo, "•")
        texto = f"[{ahora}] {icono} {mensaje}\n"
        
        # Agregar a lista de logs para exportar
        self.logs_procesamiento.append(f"[{ahora}] {icono} {mensaje}")
        
        self.txt_log.config(state='normal')
        self.txt_log.insert(tk.END, texto, tipo)
        self.txt_log.see(tk.END)
        self.txt_log.config(state='disabled')
        self.root.update_idletasks()

    def _exportar_logs(self):
        """Exporta logs a archivo txt"""
        try:
            # Crear carpeta si no existe
            if not os.path.exists(self.ruta_logs):
                os.makedirs(self.ruta_logs)
                self.add_log(f"Carpeta de logs creada: {self.ruta_logs}", "exito")
            
            # Crear archivo logs.txt
            archivo_logs = os.path.join(self.ruta_logs, "logs.txt")
            
            with open(archivo_logs, 'w', encoding='utf-8') as f:
                f.write("=" * 100 + "\n")
                f.write(f"REPORTE DE PROCESAMIENTO - DataRefresh Cupos Aprendices\n")
                f.write(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                f.write("=" * 100 + "\n\n")
                
                # Escribir estadísticas
                f.write("ESTADÍSTICAS GENERALES:\n")
                f.write("-" * 100 + "\n")
                f.write(f"Total de archivos: {self.estadisticas['total']}\n")
                f.write(f"Archivos procesados: {self.estadisticas['procesados']}\n")
                f.write(f"Exitosos: {self.estadisticas['exitosos']}\n")
                f.write(f"Fallidos: {self.estadisticas['fallidos']}\n")
                f.write(f"Tiempo total: {int(self.estadisticas['tiempo_total'])} segundos\n\n")
                
                # Detalles de cada archivo
                f.write("DETALLES DE ARCHIVOS PROCESADOS:\n")
                f.write("-" * 100 + "\n")
                for idx, detalle in enumerate(self.estadisticas['archivos_detalles'], 1):
                    f.write(f"\n[{idx}] {detalle['archivo']}\n")
                    f.write(f"    Estado: {detalle['estado']}\n")
                    f.write(f"    Duración: {detalle['duracion']} segundos\n")
                    f.write(f"    Hora: {detalle['hora']}\n")
                
                # Logs completos
                f.write("\n\n" + "=" * 100 + "\n")
                f.write("LOGS COMPLETOS:\n")
                f.write("=" * 100 + "\n")
                for log in self.logs_procesamiento:
                    f.write(log + "\n")
            
            NotificacionFlotante(self.root, "✓ Logs Exportados",
                                f"Archivo generado: {archivo_logs}\n\n"
                                f"Total de registros: {len(self.logs_procesamiento)}", "exito", 3000)
            
            self.add_log(f"Logs exportados a: {archivo_logs}", "exito")
            messagebox.showinfo("✓ Éxito", 
                              f"Archivo logs.txt generado exitosamente\n\nUbicación:\n{archivo_logs}")
            
        except Exception as e:
            error_msg = f"Error al exportar logs: {str(e)}"
            self.add_log(error_msg, "error")
            NotificacionFlotante(self.root, "✕ Error", error_msg, "error", 3000)
            messagebox.showerror("Error", error_msg)

    def _seleccionar_archivos(self):
        """Selecciona archivos"""
        archivos = filedialog.askopenfilenames(
            title="Seleccionar archivo(s) Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")],
            initialdir=self.archivo_actual if os.path.isdir(self.archivo_actual) else self.DOCUMENTOS
        )
        
        if archivos:
            self.archivos_seleccionados = list(archivos)
            cantidad = len(self.archivos_seleccionados)
            self.label_archivos_info.config(text=f"{cantidad} archivo(s) seleccionado(s)")
            self.add_log(f"Seleccionados {cantidad} archivo(s) ✓", "exito")
            NotificacionFlotante(self.root, "Archivos Seleccionados",
                                f"{cantidad} archivo(s) listos para procesar", "exito", 2000)

    def _seleccionar_carpeta(self):
        """Selecciona carpeta"""
        carpeta = filedialog.askdirectory(
            title="Seleccionar carpeta con archivos Excel",
            initialdir=self.DOCUMENTOS
        )
        
        if carpeta:
            self.archivo_actual = carpeta
            archivos_excel = [f for f in os.listdir(carpeta) if f.endswith(('.xlsx', '.xls'))]
            self.archivos_seleccionados = [os.path.join(carpeta, f) for f in archivos_excel]
            cantidad = len(self.archivos_seleccionados)
            self.label_archivos_info.config(text=f"{cantidad} archivo(s) en carpeta")
            self.add_log(f"Carpeta: {os.path.basename(carpeta)}", "importante")
            self.add_log(f"Encontrados {cantidad} archivo(s) Excel ✓", "exito")
            NotificacionFlotante(self.root, "Carpeta Seleccionada",
                                f"Se encontraron {cantidad} archivo(s) Excel", "exito", 2000)

    def _iniciar_proceso(self):
        """Inicia procesamiento"""
        if not self.archivos_seleccionados:
            NotificacionFlotante(self.root, "Advertencia",
                                "Debes seleccionar archivos primero", "alerta", 3000)
            messagebox.showwarning("Advertencia", "Debes seleccionar archivos primero")
            return
        
        self.proceso_activo = True
        self.tiempo_inicio_general = time.time()
        self.btn_iniciar.config(state="disabled")
        
        self.estadisticas = {
            'total': len(self.archivos_seleccionados),
            'procesados': 0,
            'exitosos': 0,
            'fallidos': 0,
            'tiempo_total': 0,
            'archivos_detalles': []
        }
        
        self.tree_detalles.delete(*self.tree_detalles.get_children())
        
        NotificacionFlotante(self.root, "Iniciando",
                            f"Procesando {len(self.archivos_seleccionados)} archivo(s)...", "info", 2000)
        
        hilo = threading.Thread(target=self._ejecutar_procesamiento, daemon=True)
        hilo.start()

    def _ejecutar_procesamiento(self):
        """Ejecuta procesamiento - VERSIÓN ROBUSTA"""
        excel = None
        libro = None
        
        try:
            self.add_log("=" * 100, "importante")
            self.add_log(f"INICIANDO PROCESAMIENTO - {len(self.archivos_seleccionados)} archivo(s)", "importante")
            self.add_log("=" * 100, "importante")
            
            # Inicializar Excel
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            
            # Configurar cálculo manual
            try:
                excel.Calculation = -4135  # xlCalculationManual
            except:
                pass
            
            for idx, archivo in enumerate(self.archivos_seleccionados, 1):
                # Cancelar si se solicita
                if not self.proceso_activo:
                    self.add_log("Procesamiento cancelado por usuario", "advertencia")
                    break
                
                self.archivo_actual_procesando = os.path.basename(archivo)
                self.tiempo_inicio_archivo = time.time()
                
                self.add_log("", "info")
                self.add_log(f"[{idx}/{len(self.archivos_seleccionados)}] Procesando: {self.archivo_actual_procesando}", "archivo")
                
                libro = None
                try:
                    # Abrir libro
                    self.add_log(f"   ↳ Abriendo archivo...", "info")
                    libro = excel.Workbooks.Open(archivo, UpdateLinks=False)
                    time.sleep(0.3)
                    
                    # Actualizar conexiones
                    if self.var_conexiones.get():
                        try:
                            self.add_log(f"   ↳ Actualizando conexiones...", "info")
                            libro.RefreshAll()
                            time.sleep(0.5)
                        except Exception as e:
                            self.add_log(f"   ⚠ Conexiones: {str(e)[:40]}", "advertencia")
                    
                    # Recalcular fórmulas
                    if self.var_calcular.get():
                        try:
                            self.add_log(f"   ↳ Recalculando fórmulas...", "info")
                            excel.Calculate()
                            time.sleep(0.2)
                        except Exception as e:
                            self.add_log(f"   ⚠ Cálculo: {str(e)[:40]}", "advertencia")
                    
                    # Guardar archivo
                    if self.var_guardar.get():
                        try:
                            self.add_log(f"   ↳ Guardando archivo...", "info")
                            libro.Save()
                            time.sleep(0.3)
                        except Exception as e:
                            self.add_log(f"   ⚠ Guardado: {str(e)[:40]}", "advertencia")
                    
                    duracion = time.time() - self.tiempo_inicio_archivo
                    self.add_log(f"✓ Completado en {duracion:.2f}s", "exito")
                    
                    self.estadisticas['procesados'] += 1
                    self.estadisticas['exitosos'] += 1
                    self.estadisticas['tiempo_total'] += duracion
                    
                    detalle = {
                        'archivo': os.path.basename(archivo),
                        'estado': 'Exitoso ✓',
                        'duracion': f"{duracion:.2f}",
                        'hora': datetime.now().strftime("%H:%M:%S")
                    }
                    
                except Exception as e:
                    duracion = time.time() - self.tiempo_inicio_archivo
                    error_msg = str(e)[:50]
                    self.add_log(f"✕ Error: {error_msg}", "error")
                    
                    self.estadisticas['procesados'] += 1
                    self.estadisticas['fallidos'] += 1
                    self.estadisticas['tiempo_total'] += duracion
                    
                    detalle = {
                        'archivo': os.path.basename(archivo),
                        'estado': 'Error ✕',
                        'duracion': f"{duracion:.2f}",
                        'hora': datetime.now().strftime("%H:%M:%S")
                    }
                
                finally:
                    # Cerrar libro siempre
                    try:
                        if libro:
                            libro.Close(SaveChanges=False)
                            time.sleep(0.2)
                    except:
                        pass
                
                self.estadisticas['archivos_detalles'].append(detalle)
                self._actualizar_ui_procesamiento()
            
            # Finalización
            self.add_log("", "info")
            self.add_log("=" * 100, "importante")
            
            tiempo_secs = int(self.estadisticas['tiempo_total'])
            horas = tiempo_secs // 3600
            minutos = (tiempo_secs % 3600) // 60
            segundos = tiempo_secs % 60
            
            resumen = f"COMPLETADO | ✓ {self.estadisticas['exitosos']} | ✕ {self.estadisticas['fallidos']} | ⏱️ {horas:02d}:{minutos:02d}:{segundos:02d}"
            self.add_log(resumen, "importante")
            self.add_log("=" * 100, "importante")
            
            NotificacionFlotante(self.root, "✓ Procesamiento Completado",
                                f"Total: {self.estadisticas['total']}\n"
                                f"Exitosos: {self.estadisticas['exitosos']}\n"
                                f"Fallidos: {self.estadisticas['fallidos']}\n"
                                f"Tiempo: {horas:02d}:{minutos:02d}:{segundos:02d}", "exito", 4000)
            
            if self.var_notificaciones.get():
                messagebox.showinfo("✓ Completado",
                    f"Procesamiento finalizado\n\n"
                    f"Total: {self.estadisticas['total']}\n"
                    f"Exitosos: {self.estadisticas['exitosos']}\n"
                    f"Fallidos: {self.estadisticas['fallidos']}\n"
                    f"Tiempo: {horas:02d}:{minutos:02d}:{segundos:02d}")
            
        except Exception as e:
            error_critico = f"ERROR CRÍTICO: {str(e)[:100]}"
            self.add_log(error_critico, "error")
            NotificacionFlotante(self.root, "Error Crítico", str(e)[:100], "error", 4000)
        
        finally:
            # Limpiar Excel
            try:
                if excel:
                    excel.ScreenUpdating = True
                    try:
                        excel.Calculation = -4105  # xlCalculationAutomatic
                    except:
                        pass
                    
                    try:
                        excel.Quit()
                    except:
                        pass
            except:
                pass
            
            self.btn_iniciar.config(state="normal")
            self.proceso_activo = False
            self._guardar_configuracion()

    def _actualizar_ui_procesamiento(self):
        """Actualiza UI durante procesamiento"""
        total = self.estadisticas['total']
        procesados = self.estadisticas['procesados']
        porcentaje = int((procesados / total * 100)) if total > 0 else 0
        
        # Labels de estadísticas
        self.labels_stats['total'].config(text=str(total))
        self.labels_stats['procesados'].config(text=str(procesados))
        self.labels_stats['exitosos'].config(text=str(self.estadisticas['exitosos']))
        self.labels_stats['fallidos'].config(text=str(self.estadisticas['fallidos']))
        
        # Tiempo total
        secs = int(self.estadisticas['tiempo_total'])
        h = secs // 3600
        m = (secs % 3600) // 60
        s = secs % 60
        self.labels_stats['tiempo_total'].config(text=f"{h:02d}:{m:02d}:{s:02d}")
        
        # Tiempo archivo actual
        if self.tiempo_inicio_archivo:
            duracion_archivo = time.time() - self.tiempo_inicio_archivo
            h_a = int(duracion_archivo) // 3600
            m_a = (int(duracion_archivo) % 3600) // 60
            s_a = int(duracion_archivo) % 60
            self.labels_stats['tiempo_archivo'].config(text=f"{h_a:02d}:{m_a:02d}:{s_a:02d}")
        
        # Archivo actual
        self.label_archivo_actual.config(text=self.archivo_actual_procesando[:80])
        
        # Barra de progreso
        self.barra_progreso.actualizar(porcentaje, animar=True)
        self.label_porcentaje.config(text=f"{porcentaje}%")
        
        # Tabla
        self.tree_detalles.delete(*self.tree_detalles.get_children())
        for idx, detalle in enumerate(self.estadisticas['archivos_detalles'], 1):
            archivo_display = detalle['archivo'][:50] + "..." if len(detalle['archivo']) > 50 else detalle['archivo']
            self.tree_detalles.insert("", "end", values=(
                str(idx),
                archivo_display,
                detalle['estado'],
                detalle['duracion'],
                detalle['hora']
            ))
        
        self.root.update_idletasks()

    def _iniciar_actualizacion_tiempo(self):
        """Actualiza tiempo en tiempo real"""
        def actualizar():
            if self.proceso_activo and self.tiempo_inicio_general:
                tiempo_transcurrido = int(time.time() - self.tiempo_inicio_general)
                h = tiempo_transcurrido // 3600
                m = (tiempo_transcurrido % 3600) // 60
                s = tiempo_transcurrido % 60
                
                # Actualizar sin detener el procesamiento
                try:
                    self.labels_stats['tiempo_total'].config(text=f"{h:02d}:{m:02d}:{s:02d}")
                except:
                    pass
            
            self.root.after(500, actualizar)
        
        actualizar()

    def _guardar_configuracion(self):
        """Guarda config"""
        config = {
            'documentos': self.DOCUMENTOS,
            'archivo_actual': self.archivo_actual,
            'guardar': self.var_guardar.get(),
            'calcular': self.var_calcular.get(),
            'conexiones': self.var_conexiones.get(),
            'macros': self.var_macros.get(),
            'notificaciones': self.var_notificaciones.get(),
            'abrir': self.var_abrir_archivo.get(),
        }
        
        try:
            with open('config_datarefresh_v6.json', 'w') as f:
                json.dump(config, f, indent=2)
        except:
            pass

    def cargar_configuracion(self):
        """Carga config"""
        try:
            if os.path.exists('config_datarefresh_v6.json'):
                with open('config_datarefresh_v6.json', 'r') as f:
                    config = json.load(f)
                    self.archivo_actual = config.get('archivo_actual', self.DOCUMENTOS)
        except:
            pass

    def on_closing(self):
        """Al cerrar"""
        if self.proceso_activo:
            if messagebox.askyesno("Confirmar", "¿Cancelar procesamiento en curso?"):
                self.proceso_activo = False
                self.root.destroy()
        else:
            self._guardar_configuracion()
            self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = DataRefreshProV6(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()
